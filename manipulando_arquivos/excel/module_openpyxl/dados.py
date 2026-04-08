import openpyxl
import os
from openpyxl import Workbook

PASTA = "arquivos"
CAMINHO_ARQUIVO = os.path.join(PASTA, "biblioteca_openpyxl.xlsx")
CAMPOS = ["isbn", "titulo", "autor", "genero", "ano_publicacao", "editora", "paginas", "status", "localizacao"]

def inicializar_diretorio():
    if not os.path.exists(PASTA):
        os.makedirs(PASTA)

def salvar_no_arquivo(lista_dados):
    inicializar_diretorio()
    wb = Workbook()
    ws = wb.active
    ws.title = "Biblioteca"

    ws.append(CAMPOS)

    # Escrever dados
    for livro in lista_dados:
        ws.append([livro.get(c, "") for c in CAMPOS])
    
    wb.save(CAMINHO_ARQUIVO)

def carregar_do_arquivo():
    if not os.path.exists(CAMINHO_ARQUIVO):
        return []

    wb = openpyxl.load_workbook(CAMINHO_ARQUIVO)
    ws = wb.active
    biblioteca_aux = []

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) <= 1: return [] # Só tem cabeçalho

    cabecalho = rows[0]
    for row in rows[1:]:
        livro = dict(zip(cabecalho, row))
        biblioteca_aux.append(livro)
    
    return biblioteca_aux